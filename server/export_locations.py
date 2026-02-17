from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def export_locations_xlsx(rows: list[dict], export_dir: Path) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = export_dir / f"roboard_locations_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"

    headers = ["part_number", "name", "old_location", "new_location", "note", "updated_at"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("part_number"),
            r.get("name"),
            r.get("old_location"),
            r.get("new_location"),
            r.get("note"),
            r.get("updated_at"),
        ])

    # Basic column sizing
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(40, len(h) + 10))

    wb.save(out_path)
    return out_path
