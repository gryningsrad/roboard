from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from db import get_conn
from usb import find_usb_mount
from export_wishlist import export_wishlist_xlsx

LOCAL_EXPORTS = Path("/home/pi/exports")

def _to_float(v: Any, default: float | None = None) -> float | None:
    v = _clean(v)
    if v in (None, ""):
        return default

    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()

    # Common Excel garbage cases
    if s.lower() in ("kg", "kgs", "kilogram", "kilograms"):
        return default

    try:
        return float(s)
    except ValueError:
        try:
            # Handles "1,25" → 1.25 (EU decimals)
            return float(s.replace(",", "."))
        except ValueError:
            return default

def _clean(v: Any):
    if v is None:
        return None

    # Handle Excel/py date/time objects so sqlite can bind them
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime):
        return v.isoformat(timespec="seconds")
    if isinstance(v, date):
        return v.isoformat()

    if isinstance(v, str):
        s = v.strip()
        return s[1:] if s.startswith("'") else s

    return v


def _iso(v: Any):
    # For columns that you explicitly want as date-ish strings
    if v is None:
        return None

    if isinstance(v, time):
        # time-only values shouldn't pretend to be dates; store as HH:MM:SS
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime):
        # If you only want date: return v.date().isoformat()
        return v.isoformat(timespec="seconds")
    if isinstance(v, date):
        return v.isoformat()

    s = str(v).strip()
    return s or None

def _to_int(v: Any, default: int = 0) -> int:
    v = _clean(v)
    if v in (None, ""):
        return default
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()

    # common cases: "1", "1.0", " 0 "
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return default

def _first_sheet(wb):
    # ALWAYS use the first worksheet. No name checks at all.
    if not wb.worksheets:
        raise ValueError("Workbook has no worksheets")
    ws = wb.worksheets[0]
    return ws, ws.title


def import_parts_replace_all(xlsx: Path) -> dict:
    """Imports Parts from first sheet. Before deleting parts, exports current wishlist to USB."""
    usb = find_usb_mount()
    export_dir = (usb / "spares_exports") if usb else (LOCAL_EXPORTS / "spares_exports")
    wishlist_file = export_wishlist_xlsx(export_dir)

    wb = load_workbook(xlsx, data_only=True)
    ws, sheet_name = _first_sheet(wb)

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    if "Number" not in idx:
        raise ValueError("Column 'Number' is required in Parts file (first sheet)")

    now = datetime.now().isoformat(timespec="seconds")

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM parts;")  # wishlist cascades

        attempted = 0
        inserted = 0

        for r in ws.iter_rows(min_row=2, values_only=True):
            num = _clean(r[idx["Number"]])
            if not num:
                continue

            def g(col):
                return _clean(r[idx[col]]) if col in idx else None

            # Convert weight safely
            weight = _to_float(g("Weight"))

            attempted += 1

            cur.execute(
                """
                INSERT OR IGNORE INTO parts (
                    number, name, qa_grading, maker_code, makers_reference, unit,
                    pref_vendor_code, order_status, default_location,
                    stock_class, stock_class_description, reserved,
                    price_class, asset, hm, attachments,
                    weight_unit, weight, alternative_available, ean, imported_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(num).strip(),
                    g("Name"),
                    g("QA Grading"),
                    g("Maker Code"),
                    g("Maker's Reference"),
                    g("Unit"),
                    g("Pref. Vendor Code"),
                    g("Order status"),
                    g("Default Location"),
                    g("Stock Class"),
                    g("Stock Class Description"),
                    _to_int(g("Reserved") or 0),
                    g("Price Class"),
                    g("Asset"),
                    g("HM"),
                    g("Attachments"),
                    g("Weight Unit"),
                    weight,
                    g("Alternative Available"),
                    (str(g("EAN")).strip() if g("EAN") not in (None, "") else None),  # migration_003
                    now,
                )
            )

            # rowcount is 1 if inserted, 0 if ignored
            inserted += (cur.rowcount or 0)

        conn.commit()
    finally:
        conn.close()

    return {
        "kind": "parts",
        "parts_imported": inserted,
        "rows_attempted": attempted,
        "rows_ignored_duplicates": attempted - inserted,
        "sheet_used": sheet_name,
        "usb_detected": bool(usb),
        "exported_wishlist_file": str(wishlist_file),
    }



def import_orders_replace_all(xlsx: Path) -> dict:
    """Imports Orders from first sheet. Replaces all orders."""
    wb = load_workbook(xlsx, data_only=True)
    ws, sheet_name = _first_sheet(wb)

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    if "Number" not in idx:
        raise ValueError("Column 'Number' is required in Orders file (first sheet)")

    now = datetime.now().isoformat(timespec="seconds")

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM orders;")

        count = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            num = _clean(r[idx["Number"]])
            if not num:
                continue

            def g(col):
                return _clean(r[idx[col]]) if col in idx else None

            cur.execute(
                """
                INSERT INTO orders (
                  number, title, vendor, del_address, form_type, form_status,
                  created, approved, ordered, confirmed, received,
                  service_order, details, estimate_total,
                  created_by, approved_by, ordered_by, imported_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(num),
                    g("Title"),
                    g("Vendor"),
                    g("Del. Address"),
                    g("Form Type"),
                    g("Form Status"),
                    _iso(g("Created")),
                    _iso(g("Approved")),
                    _iso(g("Ordered")),
                    _iso(g("Confirmed")),
                    _iso(g("Received")),
                    g("Service Order"),
                    g("Details"),
                    float(g("Estimate Total") or 0) if g("Estimate Total") not in (None, "") else None,
                    g("Created by"),
                    g("Approved by"),
                    g("Ordered by"),
                    now,
                ),
            )
            count += 1

        conn.commit()
    finally:
        conn.close()

    return {
        "kind": "orders",
        "orders_imported": count,
        "sheet_used": sheet_name,
    }
